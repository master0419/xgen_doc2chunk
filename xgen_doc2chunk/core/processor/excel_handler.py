# your_package/document_processor/excel_handler.py
"""
Excel Handler - Excel Document Processor (XLSX/XLS)

Main Features:
- Metadata extraction (title, author, subject, keywords, creation date, modification date, etc.)
- Text extraction (direct parsing via openpyxl/xlrd)
- Table extraction (Markdown or HTML conversion based on merged cells)
- Inline image extraction and local storage
- Chart processing (convert to table)
- Multi-sheet support

Class-based Handler:
- ExcelHandler class inherits from BaseHandler to manage config/image_processor
"""
from __future__ import annotations

import logging
import os
from typing import TYPE_CHECKING, Any, Dict, List, Optional, Set

from xgen_doc2chunk.core.processor.base_handler import BaseHandler
from xgen_doc2chunk.core.functions.img_processor import ImageProcessor
from xgen_doc2chunk.core.functions.chart_extractor import BaseChartExtractor
from xgen_doc2chunk.core.processor.excel_helper.excel_chart_extractor import ExcelChartExtractor

if TYPE_CHECKING:
    from openpyxl.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
    from xgen_doc2chunk.core.document_processor import CurrentFile
from xgen_doc2chunk.core.processor.excel_helper import (
    # Textbox
    extract_textboxes_from_xlsx,
    extract_textboxes_from_xls,
    # Table
    convert_xlsx_sheet_to_table,
    convert_xls_sheet_to_table,
    # Object Detection
    convert_xlsx_objects_to_tables,
    convert_xls_objects_to_tables,
)
from xgen_doc2chunk.core.processor.excel_helper.excel_metadata import (
    XLSXMetadataExtractor,
    XLSMetadataExtractor,
)
from xgen_doc2chunk.core.processor.excel_helper.excel_image_processor_xlsx import (
    ExcelImageProcessor,
)
from xgen_doc2chunk.core.processor.excel_helper.excel_image_processor_xls import (
    XLSImageProcessor,
)

logger = logging.getLogger("document-processor")


# ============================================================================
# ExcelHandler Class
# ============================================================================

class ExcelHandler(BaseHandler):
    """
    Excel Document Handler (XLSX/XLS)

    Inherits from BaseHandler to manage config and image_processor at instance level.

    Usage:
        handler = ExcelHandler(config=config, image_processor=image_processor)
        text = handler.extract_text(current_file)
    """

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._xlsx_metadata_extractor = None
        self._xls_metadata_extractor = None

    def _create_file_converter(self):
        """Create Excel-specific file converter."""
        from xgen_doc2chunk.core.processor.excel_helper.excel_file_converter import ExcelFileConverter
        return ExcelFileConverter()

    def _create_preprocessor(self):
        """Create Excel-specific preprocessor."""
        from xgen_doc2chunk.core.processor.excel_helper.excel_preprocessor import ExcelPreprocessor
        return ExcelPreprocessor()

    def _create_chart_extractor(self) -> BaseChartExtractor:
        """Create Excel-specific chart extractor."""
        return ExcelChartExtractor(self._chart_processor)

    def _create_metadata_extractor(self):
        """Create XLSX-specific metadata extractor (default)."""
        return XLSXMetadataExtractor()

    def _create_format_image_processor(self):
        """Create Excel-specific image processor (XLSX)."""
        return ExcelImageProcessor(
            directory_path=self._image_processor.config.directory_path,
            tag_prefix=self._image_processor.config.tag_prefix,
            tag_suffix=self._image_processor.config.tag_suffix,
            storage_backend=self._image_processor.storage_backend,
        )

    def _create_xls_image_processor(self):
        """Create XLS-specific image processor."""
        return XLSImageProcessor(
            directory_path=self._image_processor.config.directory_path,
            tag_prefix=self._image_processor.config.tag_prefix,
            tag_suffix=self._image_processor.config.tag_suffix,
            storage_backend=self._image_processor.storage_backend,
        )

    def _get_xls_metadata_extractor(self):
        """Get XLS-specific metadata extractor."""
        if self._xls_metadata_extractor is None:
            self._xls_metadata_extractor = XLSMetadataExtractor()
        return self._xls_metadata_extractor

    def extract_text(
        self,
        current_file: "CurrentFile",
        extract_metadata: bool = True,
        **kwargs
    ) -> str:
        """
        Extract text from Excel file.

        Args:
            current_file: CurrentFile dict containing file info and binary data
            extract_metadata: Whether to extract metadata
            **kwargs: Additional options

        Returns:
            Extracted text
        """
        file_path = current_file.get("file_path", "unknown")
        ext = current_file.get("file_extension", os.path.splitext(file_path)[1]).lower()
        # Normalize extension (remove leading dot if present)
        ext = ext.lstrip('.')
        self.logger.info(f"Excel processing: {file_path}, ext: {ext}")

        # Check if file content is actually HTML disguised as Excel
        file_data = current_file.get("file_data", b"")
        if self._is_html_content(file_data):
            self.logger.info(f"Detected HTML content in .{ext} file, processing as HTML table")
            return self._extract_html_as_excel(current_file)

        if ext == 'xlsx':
            return self._extract_xlsx(current_file, extract_metadata)
        elif ext == 'xls':
            return self._extract_xls(current_file, extract_metadata)
        else:
            raise ValueError(f"Unsupported Excel format: {ext}")

    def _extract_xlsx(
        self,
        current_file: "CurrentFile",
        extract_metadata: bool = True
    ) -> str:
        """XLSX file processing."""
        file_path = current_file.get("file_path", "unknown")
        self.logger.info(f"XLSX processing: {file_path}")

        try:
            # Step 1: Convert to Workbook using file_converter
            file_data = current_file.get("file_data", b"")
            wb = self.file_converter.convert(file_data, extension='xlsx')

            # Step 2: Preprocess - may transform wb in the future
            preprocessed = self.preprocess(wb)
            wb = preprocessed.clean_content  # TRUE SOURCE

            preload = self._preload_xlsx_data(current_file, wb, extract_metadata)

            result_parts = [preload["metadata_str"]] if preload["metadata_str"] else []
            processed_images: Set[str] = set()
            stats = {"charts": 0, "images": 0, "textboxes": 0}

            for sheet_name in wb.sheetnames:
                sheet_result = self._process_xlsx_sheet(
                    wb[sheet_name], sheet_name, preload, processed_images, stats
                )
                result_parts.append(sheet_result)

            remaining = self._process_remaining_charts(
                preload["chart_data_list"], preload["chart_idx"], processed_images, stats
            )
            if remaining:
                result_parts.append(remaining)

            result = "".join(result_parts)
            self.logger.info(
                f"XLSX processing completed: {len(wb.sheetnames)} sheets, "
                f"{stats['charts']} charts, {stats['images']} images"
            )
            return result

        except Exception as e:
            self.logger.error(f"Error in XLSX processing: {e}")
            import traceback
            self.logger.debug(traceback.format_exc())
            raise

    def _extract_xls(
        self,
        current_file: "CurrentFile",
        extract_metadata: bool = True
    ) -> str:
        """XLS file processing."""
        file_path = current_file.get("file_path", "unknown")
        self.logger.info(f"XLS processing: {file_path}")

        try:
            # Step 1: Convert to Workbook using file_converter
            file_data = current_file.get("file_data", b"")
            wb = self.file_converter.convert(file_data, extension='xls')

            # Step 2: Preprocess - may transform wb in the future
            preprocessed = self.preprocess(wb)
            wb = preprocessed.clean_content  # TRUE SOURCE

            result_parts = []
            stats = {"images": 0, "textboxes": 0}

            if extract_metadata:
                xls_extractor = self._get_xls_metadata_extractor()
                metadata_str = xls_extractor.extract_and_format(wb)
                if metadata_str:
                    result_parts.append(metadata_str + "\n\n")

            # Extract images grouped by sheet using XLSImageProcessor
            xls_image_processor = self._create_xls_image_processor()
            sheet_names = [wb.sheet_by_index(i).name for i in range(wb.nsheets)]
            images_by_sheet = xls_image_processor.extract_images_by_sheet(file_path, sheet_names)

            # Extract textboxes/shape texts from XLS
            textboxes_by_sheet = extract_textboxes_from_xls(file_path, sheet_names)

            for sheet_idx in range(wb.nsheets):
                ws = wb.sheet_by_index(sheet_idx)
                sheet_tag = self.create_sheet_tag(ws.name)
                result_parts.append(f"\n{sheet_tag}\n")

                # Process tables for this sheet
                table_contents = convert_xls_objects_to_tables(ws, wb)
                if table_contents:
                    for i, table_content in enumerate(table_contents, 1):
                        if len(table_contents) > 1:
                            result_parts.append(f"\n[Table {i}]\n{table_content}\n")
                        else:
                            result_parts.append(f"\n{table_content}\n")

                # Process images for this sheet
                sheet_images = images_by_sheet.get(sheet_idx, [])
                for idx, img_data in enumerate(sheet_images):
                    if img_data:
                        image_tag = xls_image_processor.save_image(img_data)
                        if image_tag:
                            result_parts.append(f"\n{image_tag}\n")
                            stats["images"] += 1

                # Process textboxes/shape texts for this sheet
                sheet_textboxes = textboxes_by_sheet.get(ws.name, [])
                for tb in sheet_textboxes:
                    if tb:
                        result_parts.append(f"\n[Textbox] {tb}\n")
                        stats["textboxes"] += 1

            result = "".join(result_parts)
            self.logger.info(f"XLS processing completed: {wb.nsheets} sheets, {stats['images']} images, {stats['textboxes']} textboxes")
            return result

        except Exception as e:
            self.logger.error(f"Error in XLS processing: {e}")
            import traceback
            self.logger.debug(traceback.format_exc())
            raise

    @staticmethod
    def _is_html_content(file_data: bytes) -> bool:
        """
        Check if file content is actually HTML (not a real Excel binary/ZIP).

        Some systems (e.g., government websites) export HTML tables with
        .xls/.xlsx extensions. Excel can open these, but xlrd/openpyxl cannot.

        Note: Valid XLS always starts with OLE magic (\\xd0\\xcf\\x11\\xe0...),
        and valid XLSX always starts with ZIP magic (PK\\x03\\x04).
        Neither can ever start with '<html' or '<!doctype', so these checks
        are safe with zero false positives.

        Files starting with '<?xml' need extra verification because
        Excel 2003 XML Spreadsheet format also starts with '<?xml' but
        is NOT HTML — it contains <Workbook> instead of <html>.
        """
        if not file_data or len(file_data) < 20:
            return False
        # Check first 1024 bytes for HTML signatures (skip BOM if present)
        header = file_data[:1024].lstrip(b'\xef\xbb\xbf').lstrip()
        header_lower = header.lower()
        # Definitive HTML signatures
        if header_lower.startswith(b'<html') or header_lower.startswith(b'<!doctype'):
            return True
        # For <?xml, verify <html> tag exists (exclude Excel 2003 XML Spreadsheet)
        if header_lower.startswith(b'<?xml'):
            return b'<html' in header_lower
        return False

    def _extract_html_as_excel(
        self,
        current_file: "CurrentFile",
    ) -> str:
        """
        Process an HTML file disguised as Excel (.xls/.xlsx).

        Parses HTML tables using BeautifulSoup and converts them to
        the same output format as regular Excel processing.

        Handles two patterns:
        - Single table with <thead>/<tbody> (standard HTML)
        - Header-only table followed by body-only table (e.g. government websites)
          → these are merged into one logical table before output
        """
        from bs4 import BeautifulSoup

        file_path = current_file.get("file_path", "unknown")
        file_data = current_file.get("file_data", b"")
        self.logger.info(f"HTML-as-Excel processing: {file_path}")

        try:
            # Decode HTML content
            text = self._decode_html_bytes(file_data)
            soup = BeautifulSoup(text, 'html.parser')

            tables = soup.find_all('table')
            if not tables:
                # No tables found - extract plain text
                body = soup.find('body')
                plain_text = body.get_text(separator='\n', strip=True) if body else soup.get_text(separator='\n', strip=True)
                return plain_text

            # Merge consecutive split tables:
            # A "header-only" table (has <thead> or only <th> cells, no <td>)
            # followed by a "body-only" table (has <tbody> or only <td> cells, no <th>)
            # are treated as a single logical table.
            logical_tables = self._merge_split_tables(tables)

            result_parts = []
            output_count = 0

            for logical_table in logical_tables:
                header_rows = logical_table['header_rows']  # list of <tr> tags
                body_rows = logical_table['body_rows']      # list of <tr> tags
                all_rows = header_rows + body_rows

                if not all_rows:
                    continue

                # Build grids separately so we know which rows are header
                header_grid = self._html_table_to_grid(header_rows) if header_rows else []
                body_grid = self._html_table_to_grid(body_rows) if body_rows else []

                if not header_grid and not body_grid:
                    continue

                has_merged = any(
                    int(cell.get('colspan', 1)) > 1 or int(cell.get('rowspan', 1)) > 1
                    for tr in all_rows
                    for cell in tr.find_all(['td', 'th'])
                )

                if has_merged:
                    table_str = self._grid_to_html_table(header_grid + body_grid)
                else:
                    table_str = self._grid_to_markdown_table(
                        header_grid, body_grid
                    )

                output_count += 1
                if len(logical_tables) > 1:
                    result_parts.append(f"\n[Table {output_count}]\n{table_str}\n")
                else:
                    result_parts.append(f"\n{table_str}\n")

            result = "".join(result_parts)
            self.logger.info(
                f"HTML-as-Excel processing completed: {output_count} tables extracted"
            )
            return result

        except Exception as e:
            self.logger.error(f"Error in HTML-as-Excel processing: {e}")
            import traceback
            self.logger.debug(traceback.format_exc())
            raise

    @staticmethod
    def _merge_split_tables(tables) -> List[Dict]:
        """
        Analyse a list of <table> tags and merge consecutive pairs where
        the first contains only header rows (<thead> / <th>) and the second
        contains only body rows (<tbody> / <td>).

        Returns a list of dicts with keys:
            'header_rows': list of <tr> BeautifulSoup tags
            'body_rows':   list of <tr> BeautifulSoup tags
        """
        def classify(table):
            """Return ('header', rows) | ('body', rows) | ('mixed', rows)."""
            all_tr = table.find_all('tr')
            if not all_tr:
                return 'empty', []
            has_th = bool(table.find('th'))
            has_td = bool(table.find('td'))
            thead_rows = [tr for tr in table.find_all('tr') if tr.find_parent('thead')]
            tbody_rows = [tr for tr in table.find_all('tr') if tr.find_parent('tbody')]
            # Header-only: has <thead> but no <tbody>, or only <th> cells
            if has_th and not has_td:
                return 'header_only', all_tr
            # Body-only: has <tbody> but no <thead>, or only <td> cells
            if has_td and not has_th:
                return 'body_only', all_tr
            # Mixed: single table with both <thead> and <tbody>
            return 'mixed', (thead_rows or all_tr[:1], tbody_rows or all_tr[1:])

        logical = []
        i = 0
        while i < len(tables):
            kind, rows = classify(tables[i])
            if kind == 'empty':
                i += 1
                continue
            if kind == 'header_only' and i + 1 < len(tables):
                next_kind, next_rows = classify(tables[i + 1])
                if next_kind == 'body_only':
                    logical.append({'header_rows': rows, 'body_rows': next_rows})
                    i += 2
                    continue
            if kind == 'mixed':
                header_rows, body_rows = rows
                logical.append({'header_rows': header_rows, 'body_rows': body_rows})
            elif kind == 'header_only':
                logical.append({'header_rows': rows, 'body_rows': []})
            else:  # body_only or unmerged
                logical.append({'header_rows': [], 'body_rows': rows})
            i += 1
        return logical

    @staticmethod
    def _decode_html_bytes(file_data: bytes) -> str:
        """Decode HTML bytes to string with encoding fallback."""
        for enc in ['utf-8', 'utf-8-sig', 'cp949', 'euc-kr', 'latin-1']:
            try:
                return file_data.decode(enc)
            except UnicodeDecodeError:
                continue
        return file_data.decode('utf-8', errors='replace')

    @staticmethod
    def _html_table_to_grid(rows) -> List[List[str]]:
        """Convert HTML table rows to a 2D grid, handling colspan/rowspan."""
        grid: List[List[Optional[str]]] = []
        for row_idx, row in enumerate(rows):
            cells = row.find_all(['td', 'th'])
            while len(grid) <= row_idx:
                grid.append([])
            col_idx = 0
            for cell in cells:
                # Skip columns already filled by rowspan
                while col_idx < len(grid[row_idx]) and grid[row_idx][col_idx] is not None:
                    col_idx += 1

                cell_text = cell.get_text(separator=' ', strip=True)
                colspan = int(cell.get('colspan', 1))
                rowspan = int(cell.get('rowspan', 1))

                for r in range(rowspan):
                    target_row = row_idx + r
                    while len(grid) <= target_row:
                        grid.append([])
                    while len(grid[target_row]) < col_idx + colspan:
                        grid[target_row].append(None)
                    for c in range(colspan):
                        grid[target_row][col_idx + c] = cell_text

                col_idx += colspan

        # Normalize row lengths
        max_cols = max((len(r) for r in grid), default=0)
        for row in grid:
            while len(row) < max_cols:
                row.append('')
            for i in range(len(row)):
                if row[i] is None:
                    row[i] = ''

        return grid

    @staticmethod
    def _html_table_has_merged_cells(table) -> bool:
        """Check if an HTML table has any merged cells."""
        for cell in table.find_all(['td', 'th']):
            if int(cell.get('colspan', 1)) > 1 or int(cell.get('rowspan', 1)) > 1:
                return True
        return False

    @staticmethod
    def _grid_to_markdown_table(
        header_grid: List[List[str]],
        body_grid: List[List[str]]
    ) -> str:
        """
        Convert header and body grids to a Markdown table string.

        - header_grid rows become the header line(s); only the first header
          row is used as the Markdown header (Markdown supports one header row).
        - body_grid rows become data rows.
        - If there is no header_grid, the first body row is used as the header.
        """
        if not header_grid and not body_grid:
            return ''

        if header_grid:
            header = header_grid[0]
            # Any extra header rows are prepended as data rows
            data_rows = header_grid[1:] + body_grid
        else:
            # No explicit header: promote first body row
            header = body_grid[0]
            data_rows = body_grid[1:]

        col_count = len(header)
        lines = []
        lines.append('| ' + ' | '.join(header) + ' |')
        lines.append('| ' + ' | '.join(['---'] * col_count) + ' |')
        for row in data_rows:
            lines.append('| ' + ' | '.join(row) + ' |')
        return '\n'.join(lines)

    @staticmethod
    def _grid_to_html_table(grid: List[List[str]]) -> str:
        """Convert 2D grid to HTML table string."""
        if not grid:
            return ''
        lines = ['<table>']
        for row in grid:
            lines.append('  <tr>')
            for cell in row:
                lines.append(f'    <td>{cell}</td>')
            lines.append('  </tr>')
        lines.append('</table>')
        return '\n'.join(lines)

    def _preload_xlsx_data(
        self, current_file: "CurrentFile", wb, extract_metadata: bool
    ) -> Dict[str, Any]:
        """Extract preprocessing data from XLSX file."""
        file_path = current_file.get("file_path", "unknown")
        file_stream = self.get_file_stream(current_file)

        result = {
            "metadata_str": "",
            "chart_data_list": [],  # ChartData instances from extractor
            "images_data": [],
            "textboxes_by_sheet": {},
            "chart_idx": 0,
        }

        if extract_metadata:
            result["metadata_str"] = self.extract_and_format_metadata(wb)
            if result["metadata_str"]:
                result["metadata_str"] += "\n\n"

        # Use ChartExtractor for chart extraction
        result["chart_data_list"] = self.chart_extractor.extract_all_from_file(file_stream)

        # Use format_image_processor directly for image extraction
        image_processor = self.format_image_processor
        if hasattr(image_processor, 'extract_images_from_xlsx'):
            result["images_data"] = image_processor.extract_images_from_xlsx(file_path)
        else:
            result["images_data"] = {}
        result["textboxes_by_sheet"] = extract_textboxes_from_xlsx(file_path)

        return result

    def _process_xlsx_sheet(
        self, ws, sheet_name: str, preload: Dict[str, Any],
        processed_images: Set[str], stats: Dict[str, int]
    ) -> str:
        """Process a single XLSX sheet."""
        sheet_tag = self.create_sheet_tag(sheet_name)
        parts = [f"\n{sheet_tag}\n"]

        table_contents = convert_xlsx_objects_to_tables(ws)
        if table_contents:
            for i, table_content in enumerate(table_contents, 1):
                if len(table_contents) > 1:
                    parts.append(f"\n[Table {i}]\n{table_content}\n")
                else:
                    parts.append(f"\n{table_content}\n")

        # Chart processing using ChartExtractor
        if hasattr(ws, '_charts') and ws._charts:
            chart_data_list = preload["chart_data_list"]
            for chart in ws._charts:
                if preload["chart_idx"] < len(chart_data_list):
                    chart_data = chart_data_list[preload["chart_idx"]]
                    # chart_data is already ChartData instance, format it
                    chart_output = self._format_chart_data(chart_data)
                    if chart_output:
                        parts.append(f"\n{chart_output}\n")
                        stats["charts"] += 1
                    preload["chart_idx"] += 1

        # Image processing - use format_image_processor directly
        image_processor = self.format_image_processor
        if hasattr(image_processor, 'get_sheet_images'):
            sheet_images = image_processor.get_sheet_images(ws, preload["images_data"], "")
        else:
            sheet_images = []
        for image_data, anchor in sheet_images:
            if image_data:
                image_tag = self.format_image_processor.save_image(image_data)
                if image_tag:
                    parts.append(f"\n{image_tag}\n")
                    stats["images"] += 1

        # Textbox processing
        textboxes = preload["textboxes_by_sheet"].get(sheet_name, [])
        for tb in textboxes:
            if tb:
                parts.append(f"\n[Textbox] {tb}\n")
                stats["textboxes"] += 1

        return "".join(parts)

    def _format_chart_data(self, chart_data) -> str:
        """Format ChartData using ChartProcessor."""
        from xgen_doc2chunk.core.functions.chart_extractor import ChartData

        if not isinstance(chart_data, ChartData):
            return ""

        if chart_data.has_data():
            return self.chart_processor.format_chart_data(
                chart_type=chart_data.chart_type,
                title=chart_data.title,
                categories=chart_data.categories,
                series=chart_data.series
            )
        else:
            return self.chart_processor.format_chart_fallback(
                chart_type=chart_data.chart_type,
                title=chart_data.title
            )

    def _process_remaining_charts(
        self, chart_data_list: List, chart_idx: int,
        processed_images: Set[str], stats: Dict[str, int]
    ) -> str:
        """Process remaining charts not associated with sheets."""
        parts = []
        while chart_idx < len(chart_data_list):
            chart_data = chart_data_list[chart_idx]
            chart_output = self._format_chart_data(chart_data)
            if chart_output:
                parts.append(f"\n{chart_output}\n")
                stats["charts"] += 1
            chart_idx += 1
        return "".join(parts)


__all__ = ["ExcelHandler"]
